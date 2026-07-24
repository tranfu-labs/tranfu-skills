#!/usr/bin/env python3
"""Run synthetic, offline tests for all supported collection platforms."""

from __future__ import annotations

import argparse
import copy
import csv
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


PLATFORMS = ["wechat", "xiaohongshu", "zhihu", "toutiao", "weibo"]
BATCH_ID = "selftest-batch"
REQUIRED_METRICS = {
    "wechat": {"followers_total", "followers_new", "followers_lost", "views", "posts", "shares", "saves"},
    "xiaohongshu": {"impressions", "views"},
    "zhihu": {"views", "plays", "likes", "comments", "shares", "saves"},
    "toutiao": {"followers_total", "followers_new", "followers_lost", "impressions", "views", "likes", "comments"},
    "weibo": {"followers_total", "views", "posts", "engagement", "reposts", "comments", "likes"},
}
SOURCE_URLS = {
    "wechat": "https://mp.weixin.qq.com/analytics",
    "xiaohongshu": "https://creator.xiaohongshu.com/statistics/account/v2",
    "zhihu": "https://www.zhihu.com/organization/analytics",
    "toutiao": "https://mp.toutiao.com/profile_v4/analysis",
    "weibo": "https://e.weibo.com/v1/eps/manage/home",
}
CONTENT_URLS = {
    "wechat": "https://mp.weixin.qq.com/s/synthetic",
    "xiaohongshu": "https://www.xiaohongshu.com/explore/synthetic",
    "zhihu": "https://www.zhihu.com/question/1/answer/2",
    "toutiao": "https://www.toutiao.com/article/synthetic",
    "weibo": "https://weibo.com/synthetic",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dashboard-root", type=Path)
    return parser.parse_args()


def manifest_for(platform: str, run_id: str = BATCH_ID) -> dict:
    all_rows = [
        {
            "date": "2026-01-01",
            "followers_total": 10,
            "followers_new": 1,
            "followers_lost": 0,
            "impressions": 100,
            "views": 20,
            "plays": 0,
            "posts": 1,
            "engagement": 3,
            "reposts": 0,
            "comments": 1,
            "likes": 2,
            "shares": 0,
            "saves": 1,
        },
        {
            "date": "2026-01-02",
            "followers_total": 11,
            "followers_new": 1,
            "followers_lost": 0,
            "impressions": 120,
            "views": 25,
            "plays": 0,
            "posts": 1,
            "engagement": 4,
            "reposts": 1,
            "comments": 1,
            "likes": 2,
            "shares": 1,
            "saves": 2,
        },
        {
            "date": "2026-01-03",
            "followers_total": 11,
            "followers_new": 0,
            "followers_lost": 0,
            "impressions": 80,
            "views": 15,
            "plays": 0,
            "posts": 0,
            "engagement": 1,
            "reposts": 0,
            "comments": 0,
            "likes": 1,
            "shares": 0,
            "saves": 0,
        },
    ]
    raw_rows = [
        {
            "阅读": row["views"],
            "播放": row["plays"],
            "点赞": row["likes"],
            "喜欢": 0,
            "跳转阅读原文": row["shares"],
            "粉丝展现量": [5, 6, 3][index],
            "粉丝阅读(播放)量": [1, 2, 0][index],
            "活跃粉丝数": [8, 9, 9][index],
            "总铁粉数": 1,
        }
        for index, row in enumerate(all_rows)
    ]
    daily = []
    for row, raw in zip(all_rows, raw_rows):
        daily.append({"date": row["date"], **{name: row[name] for name in REQUIRED_METRICS[platform]}, "raw": raw})

    contents = [
        {
            "content_id": f"{platform}-content-1",
            "title": f"Synthetic {platform} content",
            "publish_time": "2026-01-02T12:30:00+08:00",
            "url": CONTENT_URLS[platform],
            "format": "article",
            "impressions": 120,
            "views": 25,
            "likes": 2,
            "comments": 1,
            "shares": 1,
            "saves": 2,
            "followers_new": 1,
            "raw": {"转评赞数": 4, "点击数": 3},
        }
    ]
    sources = {}
    for metric in REQUIRED_METRICS[platform]:
        derived = platform == "weibo" and metric == "followers_total"
        sources[metric] = {
            "source_url": SOURCE_URLS[platform],
            "method": "derived" if derived else "dom_table",
            "confidence": "derived" if derived else "exact",
            "note": "Synthetic offline fixture",
        }
    sources["contents"] = {
        "source_url": SOURCE_URLS[platform],
        "method": "dom_table",
        "confidence": "exact",
        "note": "One unique synthetic content row; end of list recorded.",
    }
    summary = {}
    for metric in REQUIRED_METRICS[platform]:
        summary[metric] = daily[-1][metric] if metric == "followers_total" else sum(row[metric] for row in daily)

    return {
        "schema_version": 1,
        "run": {
            "id": run_id,
            "platform": platform,
            "account_name": f"Synthetic {platform} account",
            "account_id": f"public-{platform}-id",
            "captured_at": "2026-01-04T10:00:00+08:00",
            "range": {"start": "2026-01-01", "end": "2026-01-03", "preset": "custom"},
            "status": "success",
        },
        "account_daily": daily,
        "contents": contents,
        "series_sources": sources,
        "summary": summary,
        "limitations": ["Synthetic offline fixture; no live account data."],
    }


def run(command: list[str], cwd: Path | None = None, expect_success: bool = True) -> subprocess.CompletedProcess:
    result = subprocess.run(command, cwd=cwd, text=True, capture_output=True)
    if expect_success and result.returncode != 0:
        raise AssertionError(f"Command failed: {' '.join(command)}\n{result.stdout}\n{result.stderr}")
    if not expect_success and result.returncode == 0:
        raise AssertionError(f"Command unexpectedly succeeded: {' '.join(command)}")
    return result


def write_manifest(directory: Path, name: str, manifest: dict) -> Path:
    path = directory / f"{name}.json"
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def expect_rejected(exporter: Path, generic_root: Path, manifest_dir: Path, name: str, manifest: dict, message: str) -> None:
    path = write_manifest(manifest_dir, name, manifest)
    rejected = run(
        [sys.executable, str(exporter), "--manifest", str(path), "--output-root", str(generic_root)],
        expect_success=False,
    )
    assert message in rejected.stderr, rejected.stderr
    assert not (generic_root / manifest["run"]["id"] / manifest["run"]["platform"]).exists()


def check_generic(target: Path, expected_derived: list[str]) -> None:
    expected = {"collection.json", "account-daily.csv", "contents.csv", "collection.xlsx", "collection-report.json"}
    assert {path.name for path in target.iterdir()} == expected
    report = json.loads((target / "collection-report.json").read_text(encoding="utf-8"))
    assert report["account_daily_rows"] == 3
    assert report["content_rows"] == 1
    assert report["coverage"]["daily_dates_complete"] is True
    assert report["coverage"]["unique_content_rows"] == 1
    assert len(report["files"]) == 4
    assert report["derived_series"] == expected_derived
    workbook = load_workbook(target / "collection.xlsx", read_only=True)
    assert workbook.sheetnames == ["账号汇总", "每日趋势", "内容明细", "说明"]
    assert workbook["每日趋势"].max_row == 4


def check_formula_safety(exporter: Path, generic_root: Path, manifest_dir: Path) -> None:
    manifest = manifest_for("weibo", "selftest-formula")
    manifest["run"]["account_name"] = " \t=1+1"
    manifest["contents"][0]["title"] = "=1+1"
    manifest_path = write_manifest(manifest_dir, "formula", manifest)
    run([sys.executable, str(exporter), "--manifest", str(manifest_path), "--output-root", str(generic_root)])
    target = generic_root / "selftest-formula" / "weibo"
    workbook = load_workbook(target / "collection.xlsx", data_only=False)
    assert workbook["账号汇总"]["B3"].data_type == "s"
    assert workbook["账号汇总"]["B3"].value.startswith("'")
    assert workbook["内容明细"]["B2"].data_type == "s"
    assert workbook["内容明细"]["B2"].value == "'=1+1"
    with (target / "contents.csv").open(encoding="utf-8-sig", newline="") as handle:
        assert list(csv.DictReader(handle))[0]["title"] == "'=1+1"


def check_rejections(exporter: Path, generic_root: Path, manifest_dir: Path) -> int:
    cases: list[tuple[str, dict, str]] = []

    unsafe_key = manifest_for("wechat", "reject-unsafe-key")
    unsafe_key["run"]["accessToken"] = "synthetic"
    cases.append(("unsafe-key", unsafe_key, "Forbidden credential"))

    unsafe_url = manifest_for("wechat", "reject-unsafe-url")
    unsafe_url["series_sources"]["views"]["source_url"] += "?token=synthetic"
    cases.append(("unsafe-url", unsafe_url, "Authentication query parameter"))

    chinese_token = manifest_for("wechat", "reject-chinese-token")
    chinese_token["account_daily"][0]["raw"]["访问令牌"] = "synthetic"
    cases.append(("chinese-token", chinese_token, "Forbidden credential"))

    local_storage = manifest_for("wechat", "reject-local-storage")
    local_storage["account_daily"][0]["raw"]["localStorage"] = {"theme": "light"}
    cases.append(("local-storage", local_storage, "Forbidden credential"))

    html_artifact = manifest_for("wechat", "reject-html")
    html_artifact["account_daily"][0]["raw"]["html"] = "<main>synthetic</main>"
    cases.append(("html", html_artifact, "Forbidden credential"))

    request_headers = manifest_for("wechat", "reject-request-headers")
    request_headers["account_daily"][0]["raw"]["requestHeaders"] = {"Accept": "text/html"}
    cases.append(("request-headers", request_headers, "Forbidden credential"))

    request_body = manifest_for("wechat", "reject-request-body")
    request_body["account_daily"][0]["raw"]["requestBody"] = "synthetic"
    cases.append(("request-body", request_body, "Forbidden credential"))

    response_body = manifest_for("wechat", "reject-response-body")
    response_body["account_daily"][0]["raw"]["responseBody"] = "synthetic"
    cases.append(("response-body", response_body, "Forbidden credential"))

    backend_url = manifest_for("wechat", "reject-backend-url")
    backend_url["account_daily"][0]["raw"]["backgroundUrl"] = (
        "https://mp.weixin.qq.com/cgi-bin/home?action=list&page=1"
    )
    cases.append(("backend-url", backend_url, "Backend URL is forbidden"))

    duplicate_date = manifest_for("zhihu", "reject-duplicate-date")
    duplicate_date["account_daily"][1]["date"] = duplicate_date["account_daily"][0]["date"]
    cases.append(("duplicate-date", duplicate_date, "Duplicate account_daily date"))

    missing_date = manifest_for("toutiao", "reject-missing-date")
    del missing_date["account_daily"][1]
    cases.append(("missing-date", missing_date, "cover every date"))

    null_gap = manifest_for("xiaohongshu", "reject-null-gap")
    null_gap["account_daily"][1]["views"] = None
    cases.append(("null-gap", null_gap, "must not contain null date gaps"))

    missing_source = manifest_for("wechat", "reject-missing-source")
    del missing_source["series_sources"]["views"]
    cases.append(("missing-source", missing_source, "missing series_sources provenance"))

    unproven_empty_contents = manifest_for("wechat", "reject-unproven-empty-contents")
    unproven_empty_contents["contents"] = []
    del unproven_empty_contents["series_sources"]["contents"]
    cases.append(("unproven-empty-contents", unproven_empty_contents, "contents completion provenance"))

    null_summary = manifest_for("toutiao", "reject-null-summary")
    null_summary["summary"]["views"] = None
    cases.append(("null-summary", null_summary, "missing summary reconciliation"))

    mismatch = manifest_for("toutiao", "reject-mismatch")
    mismatch["summary"]["views"] = 999
    cases.append(("mismatch", mismatch, "does not match daily value"))

    negative = manifest_for("weibo", "reject-negative")
    negative["account_daily"][0]["views"] = -1
    cases.append(("negative", negative, "non-negative integer"))

    non_finite = manifest_for("wechat", "reject-nan")
    non_finite["account_daily"][0]["views"] = float("nan")
    cases.append(("nan", non_finite, "non-negative integer or null"))

    duplicate_content = manifest_for("xiaohongshu", "reject-duplicate-content")
    duplicate_content["contents"].append(copy.deepcopy(duplicate_content["contents"][0]))
    cases.append(("duplicate-content", duplicate_content, "Duplicate content row"))

    naive_time = manifest_for("zhihu", "reject-naive-time")
    naive_time["contents"][0]["publish_time"] = "2026-01-02T12:30:00"
    cases.append(("naive-time", naive_time, "must include a UTC offset"))

    unknown = manifest_for("wechat", "reject-unknown")
    unknown["account_daily"][0]["viewz"] = 20
    cases.append(("unknown", unknown, "Unsupported field"))

    partial = manifest_for("wechat", "reject-partial-without-limit")
    partial["run"]["status"] = "partial"
    partial["limitations"] = []
    cases.append(("partial", partial, "must explain the limitation"))

    missing_end_total = manifest_for("toutiao", "reject-missing-end-total")
    missing_end_total["account_daily"][-1]["followers_total"] = None
    cases.append(("missing-end-total", missing_end_total, "must not contain null date gaps"))

    zhihu_raw = manifest_for("zhihu", "reject-zhihu-raw")
    zhihu_raw["run"]["status"] = "partial"
    del zhihu_raw["account_daily"][0]["raw"]["阅读"]
    cases.append(("zhihu-raw", zhihu_raw, "required for populated Zhihu data"))

    for name, manifest, message in cases:
        expect_rejected(exporter, generic_root, manifest_dir, name, manifest, message)
    return len(cases)


def check_dashboard(
    builder: Path,
    dashboard_root: Path,
    generic_root: Path,
    raw_manifests: dict[str, Path],
    temp: Path,
) -> int:
    dashboard_output = temp / "dashboard-imports" / BATCH_ID
    raw_rejected = run(
        [
            "node", str(builder), "--manifest", str(raw_manifests["wechat"]),
            "--project-root", str(dashboard_root), "--output", str(dashboard_output),
        ],
        expect_success=False,
    )
    assert "only accept collection.json" in raw_rejected.stderr

    for platform in PLATFORMS:
        collection = generic_root / BATCH_ID / platform / "collection.json"
        result = run(
            [
                "node", str(builder), "--manifest", str(collection),
                "--project-root", str(dashboard_root), "--output", str(dashboard_output),
            ]
        )
        payload = json.loads(result.stdout)
        assert payload["validation"]
        if platform == "xiaohongshu":
            assert payload["warnings"]

    formula_collection = generic_root / "selftest-formula" / "weibo" / "collection.json"
    symlink_target = temp / "symlink-target"
    symlink_target.mkdir()
    symlink_output = temp / "symlink-batches" / "selftest-formula"
    symlink_output.parent.mkdir()
    symlink_output.symlink_to(symlink_target, target_is_directory=True)
    symlink_rejected = run([
        "node", str(builder), "--manifest", str(formula_collection),
        "--project-root", str(dashboard_root), "--output", str(symlink_output),
    ], expect_success=False)
    assert "symlink batch directory" in symlink_rejected.stderr

    nested_target = temp / "nested-symlink-target"
    nested_target.mkdir()
    nested_output = temp / "nested-symlink-batches" / "selftest-formula"
    nested_output.mkdir(parents=True)
    (nested_output / "weibo").symlink_to(nested_target, target_is_directory=True)
    nested_rejected = run([
        "node", str(builder), "--manifest", str(formula_collection),
        "--project-root", str(dashboard_root), "--output", str(nested_output),
    ], expect_success=False)
    assert "symlink directory" in nested_rejected.stderr
    assert not list(nested_target.iterdir())

    formula_output = temp / "dashboard-imports" / "selftest-formula"
    run([
        "node", str(builder), "--manifest", str(formula_collection),
        "--project-root", str(dashboard_root), "--output", str(formula_output),
    ])
    formula_file = next((formula_output / "weibo").glob("微博数据_*.xlsx"))
    formula_book = load_workbook(formula_file, data_only=False)
    assert formula_book["账号汇总"]["B3"].data_type == "s"
    assert formula_book["账号汇总"]["B3"].value.startswith("'")
    assert formula_book["单条博文"]["B2"].data_type == "s"
    assert formula_book["单条博文"]["B2"].value == "'=1+1"

    root_names = {path.name for path in dashboard_output.iterdir() if path.is_file()}
    assert "user_analysis.xls" in root_names
    assert any(name.startswith("tendency_") and name.endswith(".xls") for name in root_names)
    assert "近30日观看数据.xlsx" in root_names
    assert "笔记列表明细表.xlsx" not in root_names
    assert "日报表.xls" in root_names
    assert any(name.startswith("数据趋势_") for name in root_names)
    assert any(name.startswith("粉丝趋势_") for name in root_names)
    assert len(list((dashboard_output / "weibo").glob("微博数据_*.xlsx"))) == 1

    parser_files = sorted(str(path) for path in dashboard_output.iterdir() if path.is_file())
    parser_code = """
import path from "node:path";
import { parseSocialFile } from "./lib/import/parsers.ts";
const files = JSON.parse(process.argv[1]);
const result = files.map((file) => {
  const parsed = parseSocialFile(file);
  const account = parsed.accountMetrics ?? [];
  return {
    file: path.basename(file), parserKey: parsed.parserKey, platformId: parsed.platformId,
    rows: account.length,
    views: account.reduce((sum, row) => sum + (row.viewsTotal ?? 0), 0),
    impressions: account.reduce((sum, row) => sum + (row.impressionsTotal ?? 0), 0),
    followersTotal: account.at(-1)?.followersTotal ?? null
  };
});
console.log(JSON.stringify(result));
"""
    parsed = run(
        ["node", "--import", "tsx", "--input-type=module", "-e", parser_code, json.dumps(parser_files)],
        cwd=dashboard_root,
    )
    parser_results = json.loads(parsed.stdout.strip())
    expected_rows = {
        "wechat_user_growth_html_xls": 3,
        "wechat_tendency_ole_xls": 3,
        "toutiao_data_trend_xlsx": 3,
        "toutiao_follower_trend_xlsx": 3,
        "zhihu_daily_csv_xls": 3,
        "xhs_30day_watch_xlsx": 3,
    }
    assert {item["parserKey"]: item["rows"] for item in parser_results} == expected_rows
    assert {item["platformId"] for item in parser_results} == {"wechat", "xiaohongshu", "zhihu", "toutiao"}
    by_parser = {item["parserKey"]: item for item in parser_results}
    assert by_parser["wechat_tendency_ole_xls"]["views"] == 60
    assert by_parser["toutiao_data_trend_xlsx"]["impressions"] == 300
    assert by_parser["toutiao_follower_trend_xlsx"]["followersTotal"] == 11
    assert by_parser["zhihu_daily_csv_xls"]["views"] == 60
    assert by_parser["xhs_30day_watch_xlsx"]["impressions"] == 300
    return len(parser_results)


def main() -> int:
    args = parse_args()
    base = Path(__file__).resolve().parent
    exporter = base / "export_collection.py"
    dashboard_builder = base / "build_dashboard_exports.mjs"

    with tempfile.TemporaryDirectory(prefix="social-media-login-collector-") as temp_value:
        temp = Path(temp_value)
        manifest_dir = temp / "manifests"
        generic_root = temp / "generic"
        manifest_dir.mkdir()

        raw_manifests = {}
        for platform in PLATFORMS:
            manifest_path = write_manifest(manifest_dir, platform, manifest_for(platform))
            raw_manifests[platform] = manifest_path
            run([sys.executable, str(exporter), "--manifest", str(manifest_path), "--output-root", str(generic_root)])
            check_generic(
                generic_root / BATCH_ID / platform,
                ["followers_total"] if platform == "weibo" else [],
            )

        check_formula_safety(exporter, generic_root, manifest_dir)
        rejection_count = check_rejections(exporter, generic_root, manifest_dir)
        parser_count = 0
        if args.dashboard_root:
            parser_count = check_dashboard(
                dashboard_builder,
                args.dashboard_root.resolve(),
                generic_root,
                raw_manifests,
                temp,
            )

        print(json.dumps({
            "status": "success",
            "platforms": PLATFORMS,
            "generic_exports": 5,
            "negative_cases": rejection_count,
            "formula_injection_blocked": True,
            "dashboard_exports": bool(args.dashboard_root),
            "dashboard_parsers": parser_count,
        }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
