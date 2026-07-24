#!/usr/bin/env python3
"""Validate a collection manifest and emit portable reports."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlsplit

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError as exc:  # pragma: no cover - depends on the Codex runtime
    raise SystemExit(
        "openpyxl is required. Use codex_app__load_workspace_dependencies and run this "
        "script with the bundled Python executable."
    ) from exc


PLATFORMS = {"wechat", "xiaohongshu", "zhihu", "toutiao", "weibo"}
STATUSES = {"success", "partial", "failed"}
METHODS = {"native_export", "dom_table", "svg_chart", "canvas_tooltip", "derived"}
CONFIDENCE = {"exact", "derived"}
METRICS = [
    "followers_total",
    "followers_new",
    "followers_lost",
    "impressions",
    "views",
    "plays",
    "posts",
    "engagement",
    "reposts",
    "comments",
    "likes",
    "shares",
    "saves",
]
CONTENT_METRICS = ["impressions", "views", "likes", "comments", "shares", "saves", "followers_new"]
PLATFORM_REQUIRED_METRICS = {
    "wechat": {"followers_total", "followers_new", "followers_lost", "views", "posts", "shares", "saves"},
    "xiaohongshu": {"impressions", "views"},
    "zhihu": {"views", "plays", "likes", "comments", "shares", "saves"},
    "toutiao": {"followers_total", "followers_new", "followers_lost", "impressions", "views", "likes", "comments"},
    "weibo": {"followers_total", "views", "posts", "engagement", "reposts", "comments", "likes"},
}
SOURCE_HOSTS = {
    "wechat": {"mp.weixin.qq.com"},
    "xiaohongshu": {"creator.xiaohongshu.com"},
    "zhihu": {"www.zhihu.com"},
    "toutiao": {"mp.toutiao.com"},
    "weibo": {"e.weibo.com", "dss.sc.weibo.com"},
}
PUBLIC_CONTENT_HOSTS = {
    "wechat": {"mp.weixin.qq.com"},
    "xiaohongshu": {"www.xiaohongshu.com", "xhslink.com"},
    "zhihu": {"www.zhihu.com", "zhuanlan.zhihu.com"},
    "toutiao": {"www.toutiao.com", "m.toutiao.com"},
    "weibo": {"weibo.com", "www.weibo.com", "m.weibo.cn"},
}
ANALYTICS_HOSTS = {host for hosts in SOURCE_HOSTS.values() for host in hosts}
TOP_LEVEL_FIELDS = {"schema_version", "run", "account_daily", "contents", "series_sources", "summary", "limitations"}
RUN_FIELDS = {"id", "platform", "account_name", "account_id", "captured_at", "range", "status"}
RANGE_FIELDS = {"start", "end", "preset"}
DAILY_FIELDS = {"date", "raw", *METRICS}
CONTENT_FIELDS = {"content_id", "title", "publish_time", "url", "format", "raw", *CONTENT_METRICS}
SOURCE_FIELDS = {"source_url", "method", "confidence", "note"}
SOURCE_NAMES = {*METRICS, "contents"}
SENSITIVE_KEY_SEGMENTS = {
    "password", "passwd", "cookie", "cookies", "token", "authorization", "authentication", "auth",
    "otp", "captcha", "verification", "phone", "mobile", "credential", "credentials", "secret", "session",
}
SENSITIVE_KEY_PHRASES = {
    "api_key", "pass_ticket", "local_storage", "session_storage", "request_headers", "request_body",
    "response_headers", "response_body", "network_response", "iframe_url", "page_source",
}
SENSITIVE_CHINESE = {"密码", "验证码", "手机号", "认证", "授权", "密钥", "令牌", "凭证", "会话", "登录态"}
PHONE_PATTERN = re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")
BEARER_PATTERN = re.compile(r"\b(?:Bearer|Cookie|Authorization)\s*[:=]\s*\S+", re.I)
RUN_ID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,79}$")
DATETIME_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}T")
DANGEROUS_CELL_PATTERN = re.compile(r"^[\s\x00-\x1f]*[=+@-]")
ILLEGAL_XML_PATTERN = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


class ManifestError(ValueError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", required=True, type=Path)
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path.cwd() / "output" / "social-collections",
    )
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def read_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ManifestError(f"Cannot read manifest {path}: {exc}") from exc
    if not isinstance(value, dict):
        raise ManifestError("Manifest root must be an object")
    return value


def normalize_key(value: str) -> str:
    separated = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", value.strip())
    return re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+", "_", separated).strip("_").lower()


def is_sensitive_key(value: str) -> bool:
    normalized = normalize_key(value)
    segments = set(normalized.split("_"))
    return bool(
        segments & (SENSITIVE_KEY_SEGMENTS | {"dom", "html"})
        or any(phrase in normalized for phrase in SENSITIVE_KEY_PHRASES)
        or any(term in normalized for term in SENSITIVE_CHINESE)
    )


def scan_url_auth_material(value: str, path: str) -> None:
    if "://" not in value:
        return
    try:
        parsed = urlsplit(value)
    except ValueError as exc:
        raise ManifestError(f"Malformed URL at {path}") from exc
    if parsed.username is not None or parsed.password is not None:
        raise ManifestError(f"URL userinfo is forbidden at {path}")
    for key, _ in parse_qsl(parsed.query, keep_blank_values=True):
        if is_sensitive_key(key):
            raise ManifestError(f"Authentication query parameter is forbidden at {path}")
    allowed_url_field = re.fullmatch(r"\$\.series_sources\.[^.]+\.source_url", path) or re.fullmatch(
        r"\$\.contents\[\d+\]\.url", path
    )
    if parsed.hostname in ANALYTICS_HOSTS and not allowed_url_field:
        raise ManifestError(f"Backend URL is forbidden outside a defined URL field at {path}")


def scan_sensitive(value: Any, path: str = "$") -> None:
    if isinstance(value, dict):
        for key, child in value.items():
            if is_sensitive_key(str(key)):
                raise ManifestError(f"Forbidden credential or private field at {path}.{key}")
            scan_sensitive(child, f"{path}.{key}")
    elif isinstance(value, list):
        for index, child in enumerate(value):
            scan_sensitive(child, f"{path}[{index}]")
    elif isinstance(value, str):
        if PHONE_PATTERN.search(value):
            raise ManifestError(f"Possible mainland China phone number at {path}")
        if BEARER_PATTERN.search(value):
            raise ManifestError(f"Possible authentication material at {path}")
        scan_url_auth_material(value, path)


def require_object(value: Any, path: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ManifestError(f"{path} must be an object")
    return value


def require_list(value: Any, path: str) -> list[Any]:
    if not isinstance(value, list):
        raise ManifestError(f"{path} must be an array")
    return value


def require_text(value: Any, path: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ManifestError(f"{path} must be a non-empty string")
    return value.strip()


def ensure_only_keys(value: dict[str, Any], allowed: set[str], path: str) -> None:
    unknown = sorted(set(value) - allowed)
    if unknown:
        raise ManifestError(f"Unsupported field(s) at {path}: {', '.join(unknown)}")


def parse_date(value: Any, path: str) -> date:
    text = require_text(value, path)
    try:
        parsed = date.fromisoformat(text)
    except ValueError as exc:
        raise ManifestError(f"{path} must use YYYY-MM-DD") from exc
    if parsed.isoformat() != text:
        raise ManifestError(f"{path} must use YYYY-MM-DD")
    return parsed


def parse_datetime(value: Any, path: str) -> datetime:
    text = require_text(value, path)
    if not DATETIME_PATTERN.match(text):
        raise ManifestError(f"{path} must use an ISO 8601 datetime with T separator")
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ManifestError(f"{path} must be an ISO 8601 datetime") from exc
    if parsed.utcoffset() is None:
        raise ManifestError(f"{path} must include a UTC offset")
    return parsed


def check_number(value: Any, path: str) -> None:
    if value is None:
        return
    if isinstance(value, bool) or not isinstance(value, int):
        raise ManifestError(f"{path} must be a non-negative integer or null")
    if value < 0:
        raise ManifestError(f"{path} must be a non-negative integer")


def validate_source_url(value: Any, platform: str, path: str) -> str:
    text = require_text(value, path)
    try:
        parsed = urlsplit(text)
    except ValueError as exc:
        raise ManifestError(f"{path} must be a valid URL") from exc
    if parsed.scheme != "https" or parsed.hostname not in SOURCE_HOSTS[platform] or parsed.port is not None:
        raise ManifestError(f"{path} must use an official HTTPS analytics host for {platform}")
    if parsed.username is not None or parsed.password is not None or parsed.query or parsed.fragment:
        raise ManifestError(f"{path} must contain only HTTPS origin and pathname")
    return text


def validate_content_url(value: Any, platform: str, path: str) -> str:
    text = require_text(value, path)
    try:
        parsed = urlsplit(text)
    except ValueError as exc:
        raise ManifestError(f"{path} must be a valid URL") from exc
    if parsed.scheme != "https" or parsed.hostname not in PUBLIC_CONTENT_HOSTS[platform] or parsed.port is not None:
        raise ManifestError(f"{path} must use an official public HTTPS content host for {platform}")
    if parsed.username is not None or parsed.password is not None:
        raise ManifestError(f"{path} must not contain URL userinfo")
    for key, _ in parse_qsl(parsed.query, keep_blank_values=True):
        if is_sensitive_key(key):
            raise ManifestError(f"{path} contains an authentication query parameter")
    return text


def expected_dates(start: date, end: date) -> list[str]:
    return [(start + timedelta(days=offset)).isoformat() for offset in range((end - start).days + 1)]


def validate_zhihu_semantics(rows: list[dict[str, Any]], status: str) -> None:
    for index, row in enumerate(rows):
        raw = row.get("raw", {})
        values: dict[str, int] = {}
        canonical_present = any(row.get(metric) is not None for metric in ["views", "plays", "likes"])
        for name in ["阅读", "播放", "点赞", "喜欢"]:
            value = raw.get(name)
            if value is None:
                if status == "success" or canonical_present:
                    raise ManifestError(f"account_daily[{index}].raw.{name} is required for populated Zhihu data")
                continue
            check_number(value, f"account_daily[{index}].raw.{name}")
            values[name] = value
        if len(values) != 4:
            continue
        expected = {
            "views": values["阅读"] + values["播放"],
            "plays": values["播放"],
            "likes": values["点赞"] + values["喜欢"],
        }
        for metric, metric_value in expected.items():
            if row.get(metric) != metric_value:
                raise ManifestError(
                    f"account_daily[{index}].{metric} must match Zhihu raw components ({metric_value})"
                )


def populated_metrics(rows: list[dict[str, Any]]) -> set[str]:
    return {metric for metric in METRICS if any(row.get(metric) is not None for row in rows)}


def validate_success(
    platform: str,
    rows: list[dict[str, Any]],
    contents: list[dict[str, Any]],
    sources: dict[str, Any],
    summary: dict[str, Any],
    start: date,
    end: date,
    preset: str | None,
) -> None:
    required_dates = expected_dates(start, end)
    actual_dates = sorted(row["date"] for row in rows)
    if actual_dates != required_dates:
        raise ManifestError("A successful manifest must cover every date in run.range exactly once")
    if preset == "30d" and len(required_dates) != 30:
        raise ManifestError("run.range with preset=30d must contain exactly 30 inclusive dates")

    present = populated_metrics(rows)
    missing = sorted(PLATFORM_REQUIRED_METRICS[platform] - present)
    if missing:
        raise ManifestError(f"Successful {platform} data is missing required metric(s): {', '.join(missing)}")
    for metric in present:
        if any(row.get(metric) is None for row in rows):
            raise ManifestError(f"Successful metric {metric} must not contain null date gaps")
        if metric not in sources:
            raise ManifestError(f"Successful metric {metric} is missing series_sources provenance")
        if summary.get(metric) is None:
            raise ManifestError(f"Successful metric {metric} is missing summary reconciliation")
    if "contents" not in sources:
        raise ManifestError(f"Successful {platform} data must record contents completion provenance")
    if contents and "contents" not in sources:
        raise ManifestError("Content rows require series_sources.contents provenance")


def validate_manifest(manifest: dict[str, Any]) -> dict[str, Any]:
    scan_sensitive(manifest)
    ensure_only_keys(manifest, TOP_LEVEL_FIELDS, "$")
    if manifest.get("schema_version") != 1:
        raise ManifestError("schema_version must equal 1")

    run = require_object(manifest.get("run"), "run")
    ensure_only_keys(run, RUN_FIELDS, "run")
    run_id = require_text(run.get("id"), "run.id")
    if not RUN_ID_PATTERN.fullmatch(run_id):
        raise ManifestError("run.id contains unsupported path characters")
    platform = require_text(run.get("platform"), "run.platform")
    if platform not in PLATFORMS:
        raise ManifestError(f"Unsupported platform: {platform}")
    require_text(run.get("account_name"), "run.account_name")
    if run.get("account_id") is not None:
        require_text(run.get("account_id"), "run.account_id")
    parse_datetime(run.get("captured_at"), "run.captured_at")
    status = require_text(run.get("status"), "run.status")
    if status not in STATUSES:
        raise ManifestError(f"Unsupported run.status: {status}")

    range_value = require_object(run.get("range"), "run.range")
    ensure_only_keys(range_value, RANGE_FIELDS, "run.range")
    start = parse_date(range_value.get("start"), "run.range.start")
    end = parse_date(range_value.get("end"), "run.range.end")
    if start > end:
        raise ManifestError("run.range.start must not be after run.range.end")
    preset = None
    if range_value.get("preset") is not None:
        preset = require_text(range_value.get("preset"), "run.range.preset")

    rows = require_list(manifest.get("account_daily"), "account_daily")
    seen_dates: set[str] = set()
    for index, row_value in enumerate(rows):
        row = require_object(row_value, f"account_daily[{index}]")
        ensure_only_keys(row, DAILY_FIELDS, f"account_daily[{index}]")
        row_date = parse_date(row.get("date"), f"account_daily[{index}].date")
        if not start <= row_date <= end:
            raise ManifestError(f"account_daily[{index}].date is outside run.range")
        date_text = row_date.isoformat()
        if date_text in seen_dates:
            raise ManifestError(f"Duplicate account_daily date: {date_text}")
        seen_dates.add(date_text)
        for metric in METRICS:
            check_number(row.get(metric), f"account_daily[{index}].{metric}")
        if "raw" in row:
            require_object(row["raw"], f"account_daily[{index}].raw")

    contents = require_list(manifest.get("contents"), "contents")
    seen_content: set[tuple[str, ...]] = set()
    for index, value in enumerate(contents):
        content = require_object(value, f"contents[{index}]")
        ensure_only_keys(content, CONTENT_FIELDS, f"contents[{index}]")
        title = require_text(content.get("title"), f"contents[{index}].title")
        publish_time = require_text(content.get("publish_time"), f"contents[{index}].publish_time")
        parse_datetime(publish_time, f"contents[{index}].publish_time")
        content_id = None
        if content.get("content_id") is not None:
            content_id = require_text(content.get("content_id"), f"contents[{index}].content_id")
        url = ""
        if content.get("url") is not None:
            url = validate_content_url(content.get("url"), platform, f"contents[{index}].url")
        if content.get("format") is not None:
            require_text(content.get("format"), f"contents[{index}].format")
        for metric in CONTENT_METRICS:
            check_number(content.get(metric), f"contents[{index}].{metric}")
        if "raw" in content:
            require_object(content["raw"], f"contents[{index}].raw")
        identity = ("id", content_id) if content_id else ("fallback", title, publish_time, url)
        if identity in seen_content:
            raise ManifestError(f"Duplicate content row at contents[{index}]")
        seen_content.add(identity)

    sources = require_object(manifest.get("series_sources"), "series_sources")
    unknown_sources = sorted(set(sources) - SOURCE_NAMES)
    if unknown_sources:
        raise ManifestError(f"Unsupported series_sources key(s): {', '.join(unknown_sources)}")
    for name, value in sources.items():
        source = require_object(value, f"series_sources.{name}")
        ensure_only_keys(source, SOURCE_FIELDS, f"series_sources.{name}")
        method = require_text(source.get("method"), f"series_sources.{name}.method")
        confidence = require_text(source.get("confidence"), f"series_sources.{name}.confidence")
        validate_source_url(source.get("source_url"), platform, f"series_sources.{name}.source_url")
        require_text(source.get("note"), f"series_sources.{name}.note")
        if method not in METHODS:
            raise ManifestError(f"Unsupported source method for {name}: {method}")
        if confidence not in CONFIDENCE:
            raise ManifestError(f"Unsupported confidence for {name}: {confidence}")
        if (method == "derived") != (confidence == "derived"):
            raise ManifestError(f"Source {name} must use method=derived and confidence=derived together")

    present = populated_metrics(rows)
    for metric in present:
        if metric not in sources:
            raise ManifestError(f"Metric {metric} is missing series_sources provenance")
    if contents and "contents" not in sources:
        raise ManifestError("Content rows require series_sources.contents provenance")

    summary = require_object(manifest.get("summary"), "summary")
    ensure_only_keys(summary, set(METRICS), "summary")
    for key, value in summary.items():
        check_number(value, f"summary.{key}")
    limitations = require_list(manifest.get("limitations"), "limitations")
    for index, limitation in enumerate(limitations):
        require_text(limitation, f"limitations[{index}]")
    if not rows and not contents and status != "failed":
        raise ManifestError("Manifest must include account_daily or contents data")
    if status in {"partial", "failed"} and not limitations:
        raise ManifestError(f"A {status} manifest must explain the limitation")

    if platform == "zhihu":
        validate_zhihu_semantics(rows, status)
    if status == "success":
        validate_success(platform, rows, contents, sources, summary, start, end, preset)
    reconcile_summary(rows, summary, end.isoformat())
    return manifest


def reconcile_summary(rows: list[dict[str, Any]], summary: dict[str, Any], end_date: str) -> None:
    by_date = {row["date"]: row for row in rows}
    for metric, expected in summary.items():
        if expected is None:
            continue
        if metric == "followers_total":
            end_row = by_date.get(end_date)
            if end_row is None or end_row.get(metric) is None:
                raise ManifestError("summary.followers_total requires a value on run.range.end")
            actual = end_row[metric]
        else:
            values = [row.get(metric) for row in rows]
            if not values or any(value is None for value in values):
                raise ManifestError(f"summary.{metric} requires a complete daily series")
            actual = sum(values)
        if actual != expected:
            raise ManifestError(f"summary.{metric}={expected} does not match daily value {actual}")


def spreadsheet_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_XML_PATTERN.sub("", value)
    return f"'{cleaned}" if DANGEROUS_CELL_PATTERN.match(cleaned) else cleaned


def write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )


def write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([spreadsheet_value(value) for value in headers])
        writer.writerows([[spreadsheet_value(value) for value in row] for row in rows])


def json_cell(value: Any) -> str:
    return json.dumps(value or {}, ensure_ascii=False, separators=(",", ":"), allow_nan=False)


def daily_csv_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [row.get("date"), *[row.get(metric) for metric in METRICS], json_cell(row.get("raw"))]
        for row in sorted(rows, key=lambda item: item["date"])
    ]


CONTENT_HEADERS = [
    "content_id",
    "title",
    "publish_time",
    "url",
    "format",
    "impressions",
    "views",
    "likes",
    "comments",
    "shares",
    "saves",
    "followers_new",
    "raw",
]


def content_csv_rows(contents: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            content.get("content_id"),
            content.get("title"),
            content.get("publish_time"),
            content.get("url"),
            content.get("format"),
            content.get("impressions"),
            content.get("views"),
            content.get("likes"),
            content.get("comments"),
            content.get("shares"),
            content.get("saves"),
            content.get("followers_new"),
            json_cell(content.get("raw")),
        ]
        for content in contents
    ]


def append_rows(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append([spreadsheet_value(value) for value in row])
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def set_widths(sheet: Any, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def write_workbook(path: Path, manifest: dict[str, Any]) -> None:
    run = manifest["run"]
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "账号汇总"
    summary_rows = [
        ["字段", "值"],
        ["平台", run["platform"]],
        ["账号", run["account_name"]],
        ["账号ID", run.get("account_id")],
        ["开始日期", run["range"]["start"]],
        ["结束日期", run["range"]["end"]],
        ["采集时间", run["captured_at"]],
        ["状态", run["status"]],
        *[[key, value] for key, value in manifest["summary"].items()],
    ]
    append_rows(summary_sheet, summary_rows)
    set_widths(summary_sheet, [24, 42])

    daily_sheet = workbook.create_sheet("每日趋势")
    daily_headers = ["date", *METRICS, "raw"]
    append_rows(daily_sheet, [daily_headers, *daily_csv_rows(manifest["account_daily"])])
    set_widths(daily_sheet, [14, *([16] * len(METRICS)), 48])

    content_sheet = workbook.create_sheet("内容明细")
    append_rows(content_sheet, [CONTENT_HEADERS, *content_csv_rows(manifest["contents"])])
    set_widths(content_sheet, [20, 64, 26, 48, 14, *([14] * 7), 48])

    notes_sheet = workbook.create_sheet("说明")
    note_rows = [["项目", "说明"]]
    for name, source in manifest["series_sources"].items():
        note_rows.append([f"来源:{name}", json.dumps(source, ensure_ascii=False, allow_nan=False)])
    for index, limitation in enumerate(manifest["limitations"], start=1):
        note_rows.append([f"限制:{index}", str(limitation)])
    append_rows(notes_sheet, note_rows)
    set_widths(notes_sheet, [24, 100])
    workbook.save(path)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def coverage_report(manifest: dict[str, Any]) -> dict[str, Any]:
    rows = manifest["account_daily"]
    start = date.fromisoformat(manifest["run"]["range"]["start"])
    end = date.fromisoformat(manifest["run"]["range"]["end"])
    required_dates = expected_dates(start, end)
    return {
        "expected_daily_rows": len(required_dates),
        "daily_dates_complete": sorted(row["date"] for row in rows) == required_dates,
        "metric_non_null_rows": {
            metric: sum(row.get(metric) is not None for row in rows)
            for metric in METRICS
            if any(row.get(metric) is not None for row in rows)
        },
        "unique_content_rows": len(manifest["contents"]),
    }


def commit_staging(staging: Path, target: Path, overwrite: bool) -> None:
    backup: Path | None = None
    try:
        if target.is_symlink():
            raise ManifestError(f"Refusing to replace symlink output directory: {target}")
        if target.exists():
            if any(target.iterdir()) and not overwrite:
                raise ManifestError(f"Output directory already exists and is not empty: {target}")
            backup = staging.with_name(f"{staging.name}.previous")
            target.replace(backup)
        staging.replace(target)
        if backup is not None:
            shutil.rmtree(backup)
    except Exception:
        if backup is not None and backup.exists() and not target.exists():
            backup.replace(target)
        raise


def export(manifest: dict[str, Any], output_root: Path, overwrite: bool) -> Path:
    run = manifest["run"]
    target = output_root.resolve() / run["id"] / run["platform"]
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists() and any(target.iterdir()) and not overwrite:
        raise ManifestError(f"Output directory already exists and is not empty: {target}")

    staging = Path(tempfile.mkdtemp(prefix=f".{run['platform']}-", dir=target.parent))
    try:
        collection_path = staging / "collection.json"
        daily_path = staging / "account-daily.csv"
        contents_path = staging / "contents.csv"
        workbook_path = staging / "collection.xlsx"
        report_path = staging / "collection-report.json"

        write_json(collection_path, manifest)
        write_csv(daily_path, ["date", *METRICS, "raw"], daily_csv_rows(manifest["account_daily"]))
        write_csv(contents_path, CONTENT_HEADERS, content_csv_rows(manifest["contents"]))
        write_workbook(workbook_path, manifest)

        files = [collection_path, daily_path, contents_path, workbook_path]
        report = {
            "schema_version": 1,
            "run_id": run["id"],
            "platform": run["platform"],
            "status": run["status"],
            "account_daily_rows": len(manifest["account_daily"]),
            "content_rows": len(manifest["contents"]),
            "coverage": coverage_report(manifest),
            "derived_series": sorted(
                name
                for name, source in manifest["series_sources"].items()
                if source.get("confidence") == "derived"
            ),
            "limitations": manifest["limitations"],
            "files": [
                {"name": file.name, "bytes": file.stat().st_size, "sha256": sha256(file)}
                for file in files
            ],
        }
        write_json(report_path, report)
        commit_staging(staging, target, overwrite)
    except Exception:
        if staging.exists():
            shutil.rmtree(staging)
        raise
    return target


def main() -> int:
    args = parse_args()
    try:
        manifest = validate_manifest(read_json(args.manifest.resolve()))
        target = export(manifest, args.output_root, args.overwrite)
    except (ManifestError, OSError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(json.dumps({"status": "success", "output": str(target)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
